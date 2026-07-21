from __future__ import annotations

import json
import os
import re
import subprocess
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import httpx
from sqlalchemy import text
from sqlalchemy.orm import Session


EXTRACTS_FILE_DIR = Path("/home/webserver/models/finance/extracts")

TYPE_MAP = {
    "Зарплата": "salary",
    "Отпуск": "vacation",
    "Выплата по отчёт": "report",
    "Выписка": "extract",
}

MONTH_TO_PERIOD = {
    1: "jun", 2: "feb", 3: "mar", 4: "apr", 5: "may", 6: "june",
    7: "jul", 8: "aug", 9: "sep", 10: "oct", 11: "nov", 12: "dec",
}


class ExtractAIService:
    def __init__(self, db: Session, reference_db: Session) -> None:
        self.db = db
        self.reference_db = reference_db
        self.api_key = "bvTwJKJ7WBYii13zp1OVqU7uNwetQwpW"
        self.model = "pixtral-12b-latest"

    async def process_file(
        self,
        *,
        file_bytes: bytes,
        file_name: str | None,
        content_type: str | None,
        actor_id: str | None,
    ) -> dict[str, Any]:
        file_text = self._extract_file_text(file_bytes, content_type, file_name)
        data = await self._analyze_with_mistral(file_text, file_name)

        counterparties_id = self._find_counterparty(data.get("ogrn"), data.get("ogrnip"))
        extract_type = TYPE_MAP.get(data.get("registry_type", ""), "extract")

        registry_date = None
        if data.get("registry_date"):
            try:
                registry_date = datetime.strptime(data["registry_date"], "%Y-%m-%d").date()
            except ValueError:
                pass

        period = MONTH_TO_PERIOD.get(registry_date.month) if registry_date else None
        now = datetime.now(timezone.utc)

        self.db.execute(
            text(
                """
                INSERT INTO extracts (type, date, counterparties_id, num, period, created_at, created_by)
                VALUES (:type, :date, :counterparties_id, :num, :period, :created_at, :created_by)
                """
            ),
            {
                "type": extract_type,
                "date": registry_date,
                "counterparties_id": counterparties_id,
                "num": data.get("extract_num"),
                "period": period,
                "created_at": now,
                "created_by": actor_id,
            },
        )
        result = self.db.execute(text("SELECT LAST_INSERT_ID()")).first()
        extract_id = result[0] if result else None

        for recipient in data.get("recipients", []):
            employee_id = self._find_employee(recipient.get("fio", ""))
            consider = employee_id is not None
            item_id = str(uuid.uuid4())
            self.db.execute(
                text(
                    """
                    INSERT INTO extract_item (
                        id, extract_id, num, fio, employee_id,
                        account_num, bik, withheld, sum, result,
                        comment_result, consider
                    )
                    VALUES (
                        :id, :extract_id, :num, :fio, :employee_id,
                        :account_num, :bik, :withheld, :sum, :result,
                        :comment_result, :consider
                    )
                    """
                ),
                {
                    "id": item_id,
                    "extract_id": extract_id,
                    "num": recipient.get("num"),
                    "fio": recipient.get("fio", ""),
                    "employee_id": employee_id,
                    "account_num": recipient.get("account_num"),
                    "bik": recipient.get("bik"),
                    "withheld": recipient.get("withheld"),
                    "sum": recipient.get("sum"),
                    "result": recipient.get("result"),
                    "comment_result": recipient.get("comment_result"),
                    "consider": consider,
                },
            )

            if extract_type in ("salary", "vacation", "report"):
                self._create_operations_salary(
                    extract_type=extract_type,
                    extract_id=extract_id,
                    date=registry_date,
                    period=period,
                    employee_id=employee_id,
                    value=recipient.get("sum"),
                    result=recipient.get("result"),
                    actor_id=actor_id,
                    now=now,
                )

        self._save_file(
            file_bytes=file_bytes,
            file_name=file_name,
            content_type=content_type,
            extract_id=extract_id,
            actor_id=actor_id,
            now=now,
        )
        self.db.commit()

        return {
            "extract_id": extract_id,
            "company_name": data.get("company_name"),
            "ogrn": data.get("ogrn"),
            "ogrnip": data.get("ogrnip"),
            "registry_type": data.get("registry_type"),
            "registry_date": data.get("registry_date"),
            "extract_num": data.get("extract_num"),
            "period": period,
            "recipients_count": len(data.get("recipients", [])),
        }

    def _create_operations_salary(
        self,
        extract_type: str,
        extract_id: int | None,
        date: Any,
        period: str | None,
        employee_id: str | None,
        value: Any,
        result: str | None,
        actor_id: str | None,
        now: datetime,
    ) -> None:
        if not employee_id or not date or not period or not value:
            return
        if result != "Зачислено":
            return

        if extract_type == "salary":
            day = date.day
            type_id = 1 if 20 <= day <= 30 else 2
        elif extract_type == "report":
            type_id = 6
        elif extract_type == "vacation":
            type_id = 4
        else:
            return

        op_id = str(uuid.uuid4())
        self.db.execute(
            text(
                """
                INSERT INTO operations_salary (
                    id, date, nounth_period, year, employee_id,
                    created_by, value, type_id, method_id, extract_id, created_at
                )
                VALUES (
                    :id, :date, :nounth_period, :year, :employee_id,
                    :created_by, :value, :type_id, :method_id, :extract_id, :created_at
                )
                """
            ),
            {
                "id": op_id,
                "date": date,
                "nounth_period": period,
                "year": date.year,
                "employee_id": employee_id,
                "created_by": actor_id,
                "value": value,
                "type_id": type_id,
                "method_id": 3,
                "extract_id": extract_id,
                "created_at": now,
            },
        )

    def _save_file(
        self,
        *,
        file_bytes: bytes,
        file_name: str | None,
        content_type: str | None,
        extract_id: int | None,
        actor_id: str | None,
        now: datetime,
    ) -> None:
        file_id = str(uuid.uuid4())
        safe_name = re.sub(r"[^A-Za-zА-Яа-я0-9._-]+", "_", (file_name or "file").strip())[:180]
        storage_name = f"{file_id}_{safe_name}"
        file_path = EXTRACTS_FILE_DIR / storage_name

        EXTRACTS_FILE_DIR.mkdir(parents=True, exist_ok=True)
        with file_path.open("wb") as f:
            f.write(file_bytes)

        ext = Path(file_name).suffix if file_name else None
        self.db.execute(
            text(
                """
                INSERT INTO extract_files (
                    id, extract_id, original_name, storage_name, extension,
                    mime_type, file_path, uploaded_by, uploaded_at
                )
                VALUES (
                    :id, :extract_id, :original_name, :storage_name, :extension,
                    :mime_type, :file_path, :uploaded_by, :uploaded_at
                )
                """
            ),
            {
                "id": file_id,
                "extract_id": extract_id,
                "original_name": file_name,
                "storage_name": storage_name,
                "extension": ext,
                "mime_type": content_type,
                "file_path": str(file_path),
                "uploaded_by": actor_id,
                "uploaded_at": now,
            },
        )

    def _extract_file_text(
        self,
        file_bytes: bytes,
        content_type: str | None,
        file_name: str | None,
    ) -> str | None:
        if self._is_pdf(content_type, file_name):
            return self._extract_pdf_text(file_bytes)
        if self._is_excel(content_type, file_name):
            return self._extract_excel_text(file_bytes)
        return self._decode_text(file_bytes)

    @staticmethod
    def _is_pdf(content_type: str | None, file_name: str | None) -> bool:
        return content_type == "application/pdf" or bool(
            file_name and file_name.lower().endswith(".pdf")
        )

    @staticmethod
    def _is_excel(content_type: str | None, file_name: str | None) -> bool:
        mime_types = [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]
        extensions = (".xlsx", ".xls", ".xlsb", ".xlsm")
        return (content_type in mime_types) or bool(
            file_name and file_name.lower().endswith(extensions)
        )

    @staticmethod
    def _extract_pdf_text(file_bytes: bytes) -> str | None:
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf") as pdf_file:
                pdf_file.write(file_bytes)
                pdf_file.flush()
                result = subprocess.run(
                    ["pdftotext", "-layout", pdf_file.name, "-"],
                    capture_output=True,
                    text=True,
                    timeout=20,
                    check=False,
                )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return None
        if result.returncode != 0:
            return None
        text_value = result.stdout.strip()
        return text_value or None

    @staticmethod
    def _extract_excel_text(file_bytes: bytes) -> str | None:
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as xl_file:
                xl_file.write(file_bytes)
                tmp_path = xl_file.name
            result = subprocess.run(
                [
                    "python3", "-c",
                    f"""
import openpyxl, sys
wb = openpyxl.load_workbook('{tmp_path}', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        print('\\t'.join(str(c) if c is not None else '' for c in row))
""",
                ],
                capture_output=True,
                text=True,
                timeout=30,
                check=False,
            )
            Path(tmp_path).unlink(missing_ok=True)
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return None
        if result.returncode != 0:
            return None
        return result.stdout.strip() or None

    @staticmethod
    def _decode_text(file_bytes: bytes) -> str | None:
        for encoding in ("utf-8", "cp1251"):
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return None

    async def _analyze_with_mistral(
        self, file_text: str | None, file_name: str | None
    ) -> dict[str, Any]:
        system_prompt = (
            """
Ты анализируешь документ (реестр или выписку) и извлекаешь структурированные данные.
Верни только валидный JSON-объект без markdown и пояснений.

Схема ответа строго такая:
{
  "company_name": "Название компании",
  "ogrn": "ОГРН или null",
  "ogrnip": "ОГРНИП или null",
  "extract_num": "123 или null",
  "registry_date": "YYYY-MM-DD",
  "registry_type": "Зарплата|Отпуск|Выплата по отчёт|Выписка",
  "recipients": [
    {
      "num": 1,
      "fio": "Иванов Иван Иванович",
      "account_num": "40817810000000000001",
      "bik": "044525225",
      "withheld": 5000.00,
      "sum": 50000.00,
      "result": "Зачислено",
      "comment_result": null
    }
  ]
}

Правила:
- company_name — название компании-плательщика из документа
- ogrn — ОГРН компании (13 цифр), если есть
- ogrnip — ОГРНИП (15 цифр), если есть
- registry_date — дата реестра или выписки в формате YYYY-MM-DD
- registry_type — тип реестра: "Зарплата" если зарплатный, "Отпуск" если отпускной, "Выплата по отчёт" если выплата по отчёту, "Выписка" если обычная выписка
- recipients — массив получателей денежных средств:
  - num — порядковый номер строки
  - fio — ФИО получателя
  - account_num — номер счета
  - bik — БИК банка
  - withheld — удержанная сумма (число)
  - sum — сумма к зачислению (число)
  - result — результат зачисления
  - comment_result — комментарий или null
- Если данных нет, поле должно быть null
- Числа возвращай JSON-числами, а не строками
"""
        ).strip()

        user_text = file_name or ""
        if file_text:
            user_text = f"{user_text}\n\nСодержимое файла:\n{file_text[:30000]}"

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": [{"type": "text", "text": user_text}]},
        ]

        async with httpx.AsyncClient(timeout=120) as client:
            response = await client.post(
                "https://api.mistral.ai/v1/chat/completions",
                headers={"Authorization": f"Bearer {self.api_key}"},
                json={
                    "model": self.model,
                    "messages": messages,
                    "temperature": 0.1,
                    "response_format": {"type": "json_object"},
                },
            )
        response.raise_for_status()
        content = response.json()["choices"][0]["message"]["content"]
        return self._parse_json(content)

    @staticmethod
    def _parse_json(content: str) -> dict[str, Any]:
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", content, flags=re.S)
            if not match:
                raise ValueError("Mistral не вернул JSON")
            return json.loads(match.group(0))

    def _find_counterparty(self, ogrn: str | None, ogrnip: str | None) -> str | None:
        if ogrn:
            row = self.reference_db.execute(
                text(
                    "SELECT counterparties_id FROM details_llc WHERE ogrn = :ogrn LIMIT 1"
                ),
                {"ogrn": ogrn},
            ).first()
            if row:
                return row[0]
        if ogrnip:
            row = self.reference_db.execute(
                text(
                    "SELECT counterparty_id FROM details_ip WHERE ogrnip = :ogrnip LIMIT 1"
                ),
                {"ogrnip": ogrnip},
            ).first()
            if row:
                return row[0]
        return None

    def _find_employee(self, fio: str) -> str | None:
        parts = fio.strip().split()
        if len(parts) < 2:
            return None
        surname = parts[0]
        name = parts[1]
        patronymic = parts[2] if len(parts) > 2 else None

        if name and patronymic:
            row = self.db.execute(
                text(
                    """
                    SELECT id FROM employee
                    WHERE surname = :surname AND name = :name AND patronymic = :patronymic
                    LIMIT 1
                    """
                ),
                {"surname": surname, "name": name, "patronymic": patronymic},
            ).first()
        else:
            row = self.db.execute(
                text(
                    """
                    SELECT id FROM employee
                    WHERE surname = :surname AND name = :name
                    LIMIT 1
                    """
                ),
                {"surname": surname, "name": name},
            ).first()
        return row[0] if row else None
